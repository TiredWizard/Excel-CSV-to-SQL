from __future__ import unicode_literals
import xlrd
import codecs
import csv
from openpyxl import load_workbook


# import da excel (xlsx)
def convert_excel_xlsx(excel_path, sql_path, sheet_names, table_names):
    try:
        book = load_workbook(excel_path)
        out_file = codecs.open(sql_path, "w", "utf_8")
        for i, sheet_name in enumerate(sheet_names):
            table_name = table_names[i]
            sheet = book.active
            sheet = book[sheet_name]

            column_list = list()
            for col in range(1, sheet.max_column + 1):
                column_list.append(sheet.cell(1, col).value)

            sql_prefix = "INSERT INTO %s(%s) VALUES(" % (table_name, ",".join(column_list))
            sql_suffix = ");\n"

            for row in range(sheet.max_row):
                if row >= 1:
                    value_list = list()
                    for col in range(1, sheet.max_column + 1):
                        val = sheet.cell(row, col).value
                        if type(val) != float:
                            val = "'%s'" % (val,)
                        else:
                            val = str(val).replace(".0", "")
                        value_list.append(val)
                    sql = "%s%s%s" % (sql_prefix, ",".join(value_list), sql_suffix)
                    out_file.write(sql)
        out_file.close()
    except IOError as error:
        print("File %s non trovato" % excel_path)


# import da excel (xls)
def convert_excel_xls(excel_path, sql_path, sheet_names, table_names):
    try:
        book = xlrd.open_workbook(excel_path)
        out_file = codecs.open(sql_path, "a", "utf_8")
        for i, sheet_name in enumerate(sheet_names):
            table_name = table_names[i]
            sheet = book.sheet_by_name(sheet_name)

            column_list = list()
            for col in range(sheet.ncols):
                column_list.append(sheet.cell(0, col).value)

            sql_prefix = "INSERT INTO %s(%s) VALUES(" % (table_name, ",".join(column_list))
            sql_suffix = ");\n"

            for row in range(sheet.nrows):
                if row >= 1:
                    value_list = list()
                    for col in range(sheet.ncols):
                        val = sheet.cell(row, col).value
                        if type(val) != float:
                            val = "'%s'" % (val,)
                        else:
                            val = str(val).replace(".0", "")
                        value_list.append(val)
                    sql = "%s%s%s" % (sql_prefix, ",".join(value_list), sql_suffix)
                    out_file.write(sql)
        out_file.close()
    except IOError as error:
        print("File %s non trovato" % excel_path)


# import da csv
def convert_csv(csv_path, sql_path, table_name, column_list):
    try:
        with open(csv_path) as csvfile:
            csv_reader = csv.DictReader(csvfile)
            out_file = codecs.open(sql_path, "a", "utf_8")

            sql_prefix = "INSERT INTO %s(%s) VALUES(" % (table_name, ",".join(column_list))
            sql_suffix = ");\n"

            for row in csv_reader:
                value_list = list()
                for i in column_list:
                    value_list.append(row[i])
                sql = "%s%s%s" % (sql_prefix, ",".join(value_list), sql_suffix)
                out_file.write(sql)
        out_file.close()
    except IOError as error:
        print("File %s non trovato" % csv_path)


# pulizia del file sql
def clean_file(file_path):
    codecs.open(file_path, "w", "utf_8").close()


# creazione tabella
def create_table(sql_path):
    out_file = codecs.open(sql_path, "a", "utf_8")
    table_name = input("Inserire il nome della tabella: ")
    engine = input("Inserire l'engine da utilizzare: ")

    sql_prefix = "CREATE TABLE %s (" % table_name
    sql_suffix = ")ENGINE=%s;\n" % engine.upper()

    attributi = ""
    while True:
        attributo = input("inserire il nuovo attributo - end per completare l'inserimento: ")
        if attributo == "end":
            break
        else:
            attributi = attributi + attributo + ","

    if len(attributi) > 0:
        attributi = attributi[:-1]

    print(attributi)
    sql = "%s%s%s" % (sql_prefix, attributi, sql_suffix)
    out_file.write(sql)
    out_file.close()
